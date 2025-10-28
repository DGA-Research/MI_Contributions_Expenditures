"""
Combine CSV outputs into a single Excel workbook.

Usage:
    python csv_to_workbook.py --input csv_output --output finance_summary.xlsx

Requires `openpyxl`. Install with:
    python -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Iterable, List

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit(
        "Missing dependency openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc


INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def safe_sheet_name(name: str) -> str:
    """Coerce a filename into an Excel-compliant sheet name."""
    title = INVALID_SHEET_CHARS.sub("_", name)
    title = title.strip()
    if not title:
        title = "Sheet"
    if len(title) > 31:
        title = title[:31]
    return title


def add_csv_to_workbook(csv_path: Path, workbook: Workbook) -> None:
    """Append a worksheet to `workbook` populated with the CSV contents."""
    sheet_name = safe_sheet_name(csv_path.stem)
    if sheet_name in workbook.sheetnames:
        # Ensure uniqueness by appending a counter.
        suffix = 1
        candidate = f"{sheet_name[:29]}_{suffix}"
        while candidate in workbook.sheetnames:
            suffix += 1
            candidate = f"{sheet_name[:29]}_{suffix}"
        sheet_name = candidate

    ws = workbook.create_sheet(title=sheet_name)
    with csv_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        for row in reader:
            ws.append(row)


def collect_csv_files(directory: Path) -> List[Path]:
    return sorted(p for p in directory.iterdir() if p.suffix.lower() == ".csv")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Combine CSV files into a multi-sheet Excel workbook."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("csv_output"),
        help="Directory containing CSV files (default: %(default)s).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("finance_summary.xlsx"),
        help="Destination Excel filename (default: %(default)s).",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.input.exists():
        parser.error(f"Input directory not found: {args.input}")
    if not args.input.is_dir():
        parser.error(f"Input path is not a directory: {args.input}")

    csv_files = collect_csv_files(args.input)
    if not csv_files:
        parser.error(f"No CSV files found in {args.input}")

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for csv_file in csv_files:
        add_csv_to_workbook(csv_file, workbook)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Wrote workbook to {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
